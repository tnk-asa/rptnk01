# 情報システム室が公開している日報データからプリプレス分を集計
# 2019.07.04 By Tanikawa

import re
import os
import sys
import datetime
import requests
import openpyxl as excel
import codecs
import chardet

#Web情報取得 月次データを取得して一時保存
def web_get(obj):
    # 以前のデータを削除
    data_org = 'monthlydata_org.txt'
    data_tmp = 'monthlydata_tmp.txt'
    if os.path.exists(data_org):
        os.remove(data_org)
    if os.path.exists(data_tmp):
        os.remove(data_tmp)
    # 月次データを取得して保存
    try:
        get_url_info = requests.get(obj)
        get_url_info.raise_for_status()
        with open(data_org, 'wb') as file:
            for mdt in get_url_info.iter_content(4029):
                file.write(mdt)
    except requests.RequestException as e:
        print(e)
        return 'Err'
    # ファイルの文字コードを判定
    with open(data_org, 'rb') as f:
        char_code = chardet.detect(f.read())
        char_code2 = char_code['encoding']
    if re.compile('(UTF|utf)').search(char_code2):
        os.rename(data_org, data_tmp)
    else:
        # 文字コードを utf-8 に変換して保存
        fin = codecs.open(data_org, "r", "shift_jis")
        fout_utf = codecs.open(data_tmp, "w", "utf-8")
        for row in fin:
            fout_utf.write(row)
        fin.close()
        fout_utf.close()     

#日報タイプごとに集計
def aggr(mfile):
    month_data = open(mfile, "r", encoding='UTF-8')
    prepress = '^999 .*'
    timeall = timehanshita = timeseihan = timekousei = timesaisagyo = 0.00
    cntall = cnthanshita = cntseihan = cntkousei = cntsaisagyo = 0
    nip = []

    for line in month_data:
        po = re.match(prepress, line.replace(',',''))
        if po:
            nip = (po.group().split( ))
            #全て集計
            try:
                timeall += float(nip[5]) #作業時間
                cntall  += int(nip[6]) #作業点数
            except:
                pass
            #版下
            if nip[2] in {'H1', 'H2', 'H3', 'H4', 'P1', 'P2', 'P3', 'P4', 'T1', 'T2', 'T3', 'T4', 'T5', 'T6', 'T7', 'TM', 'N6', 'N7', 'J4', 'J5', 'J6', 'J7'}:
                try:
                    timehanshita += float(nip[5])
                    cnthanshita += int(nip[6])
                except:
                    pass
            #製版
            if nip[2] in {'S1', 'S2', 'S3', 'S4', 'AM', 'D1', 'D2', 'TM', 'G1', 'G2', 'G3', 'G4', 'G5'}:
                try:
                    timeseihan += float(nip[5])
                    cntseihan += int(nip[6])
                except:
                    pass
            #校正
            if nip[2] in {'N1', 'N2', 'N3', 'N4', 'K1', 'K2', 'K3', 'K4', 'K5', 'K6', 'K7', 'K8', 'K9', 'CK','KA', 'KB', 'KC', 'KS'}:
                try:
                    timekousei += float(nip[5])
                    cntkousei += int(nip[6])
                except:
                    pass
            #再作業
            if nip[2] in {'SS', 'RT', 'RP'}:
                try:
                    timesaisagyo += float(nip[5])
                    cntsaisagyo += int(nip[6])
                except:
                    pass
    return [timehanshita, timeseihan, timekousei, timeall, timesaisagyo], [cnthanshita, cntseihan, cntkousei, cntall, cntsaisagyo]
    month_data.close()

#########初期処理###########
#.pyがあるディレクトリをカレントディレクトリに
#os.chdir(os.path.dirname(os.path.abspath(__file__)))
#exeファイル用
#dir_test0 = os.chdir(os.path.dirname(sys.executable))
print("----- Start Processing -----")

#ディレクトリ設定
try:
    #base_path = sys._MEIPASS
    base_path = os.path.dirname(sys.argv[0])
    print("1:", base_path)
except Exception:
    base_path = os.path.dirname(__file__)
    print("2:", base_path)

#########月次実績データ取得###########
yyyy = datetime.date.today().year
if datetime.date.today().month < 4:
    yyyy -= 1
url1 = 'http://192.168.165.100/w3root/WebSTOK/ReportWeMo/'
#url1 = 'http://192.168.165.150/w3root/WebSTOK/ReportWeMo/'
#url1 = 'http://192.168.103.86/tbdoc/testdata/' # テスト用
chkurl = ''

pptable = excel.Workbook()
ws1 = pptable.active
ws1.title = '作業時間h'
ws2 = pptable.create_sheet('作業点数')
y = 1
x = 1

#if os.path.exists('monthlydata_tmp.txt'):
#    os.remove('monthlydata_tmp.txt')

for yr in range(yyyy-2, yyyy+1):
    for m in range(4, 16):
        if yr >= datetime.date.today().year and m >= datetime.date.today().month:
            break
        if m > 15:
            break
        else:
            if m > 12:
                m -= 12
            mm = (str(m).zfill(2))
            urlx = url1 + str(yr) + 'yr/month' + mm + '.TXT'
            chkurl = web_get(urlx)
            if chkurl == 'Err':
                print(urlx,'not exist')
                continue
            #実績ファイル出力
            mdata = "monthlydata_tmp.txt"
            if os.path.isfile(mdata) == 1:
                timepp = cntpp = 0
                timepp, cntpp = aggr(mdata)
                if yr == yyyy:
                    y = 1
                if yr == yyyy-1:
                    y = 9
                if yr == yyyy-2:
                    y = 17
                if 4 <= m:
                    x = m - 2
                if 4 > m:
                    x = m + 10
                if y > 17:
                    break
                else:
                    if m in {1, 4}:
                        if m == 1:
                            ws1.cell(column=x, row=y, value=yr+1)
                        elif m == 4:
                            #見出し出力
                            ws1.cell(column=x, row=y, value=yr)
                            ws1.cell(column=1, row=y+2, value='版下')
                            ws1.cell(column=1, row=y+3, value='製版')
                            ws1.cell(column=1, row=y+4, value='校正')
                            ws1.cell(column=1, row=y+5, value='工程合計')
                            ws1.cell(column=1, row=y+6, value='総合計')
                            ws1.cell(column=1, row=y+7, value='再作業')
                            ws2.cell(column=x, row=y, value=yr)
                            ws2.cell(column=1, row=y+2, value='版下')
                            ws2.cell(column=1, row=y+3, value='製版')
                            ws2.cell(column=1, row=y+4, value='校正')
                            ws2.cell(column=1, row=y+5, value='工程合計')
                            ws2.cell(column=1, row=y+6, value='総合計')
                            ws2.cell(column=1, row=y+7, value='再作業')
                    #集計値出力
                    ws1.cell(column=x, row=y+1, value=mm)
                    ws1.cell(column=x, row=y+2, value=round(timepp[0],2)) #版下
                    ws1.cell(column=x, row=y+3, value=round(timepp[1],2)) #製版
                    ws1.cell(column=x, row=y+4, value=round(timepp[2],2)) #校正
                    ws1.cell(column=x, row=y+5, value=round(timepp[0]+timepp[1]+timepp[2],2)) #工程合計
                    ws1.cell(column=x, row=y+6, value=round(timepp[3],2)) #総合計
                    ws1.cell(column=x, row=y+7, value=round(timepp[4],2)) #再作業
                    ws2.cell(column=x, row=y+1, value=mm)
                    ws2.cell(column=x, row=y+2, value=round(cntpp[0],2)) #版下
                    ws2.cell(column=x, row=y+3, value=round(cntpp[1],2)) #製版
                    ws2.cell(column=x, row=y+4, value=round(cntpp[2],2)) #校正
                    ws2.cell(column=x, row=y+5, value=round(cntpp[0]+cntpp[1]+cntpp[2],2)) #工程合計
                    ws2.cell(column=x, row=y+6, value=round(cntpp[3],2)) #総合計
                    ws2.cell(column=x, row=y+7, value=round(cntpp[4],2)) #再作業

#出力ファイル名
out_file = base_path + '\ppjisseki' + str(yyyy) + str(mm) + '.xlsx'
pptable.save(out_file)

print("----- Finished -----")